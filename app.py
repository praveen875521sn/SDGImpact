"""
ECP Impact Dashboard — Flask Web Application
============================================
Reads directly from 3 raw Excel files at startup:
  data/ECP.xlsx
  data/Attendance_data.xlsx
  data/FA.xlsx

No pre-processing needed. All data computed at startup and served from RAM.
Architecture mirrors center-dashboard/app.py
"""

from flask import Flask, render_template, jsonify, request
import openpyxl, os, re, time
from collections import defaultdict

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR  = os.path.join(BASE_DIR, "data")

# ── Allowed candidate statuses ────────────────────────────────────────────────
ALLOWED_STATUSES = {
    "assessed","attrition","batchEnd","batchProcessed","placed",
    "queuedForAssessment","queuedForPlacement","rejectedByCandidate","rejectedByCenter"
}

FA_STATUS_COL = "Pass/Fail/Attempted/Not Attempted"

# ── College normalisation ─────────────────────────────────────────────────────
def normalise_college(name):
    if not isinstance(name, str) or not name.strip():
        return "Not Specified"
    n = name.strip()
    nl = re.sub(r"\s+", " ", n.lower())
    if nl in ("not specified","n/a","na","nil","none","--","-","not applicable",""):
        return "Not Specified"
    if nl in ("other locations","other location","others","other"):
        return "Other Locations"
    rc_kw = ["ramnarayan","ramanarayan","chellaram","chellera","challaraman","rc college","govt rc","government rc"]
    if any(k in nl for k in rc_kw):
        return "RC College / Govt Ramnarayan Chellaram College"
    if "al ameen" in nl or "alameen" in nl:
        return "Al Ameen Arts Science & Commerce College"
    if "wbchse" in nl: return "WBCHSE"
    if "wbbse"  in nl: return "WBBSE"
    govt_kw = ["zphs","zp high","zilla parishad high","govt high school","govt school",
               "government school","goverment school","government high school"]
    if any(k in nl for k in govt_kw): return "Government School / High School"
    if "tilak" in nl and "college" in nl: return "Tilak College"
    if "karnataka" in nl and any(k in nl for k in ("board","examination","education","secondary")):
        return "Karnataka Board of Examination"
    if ("maharashtra" in nl and "board" in nl) or "msbshse" in nl:
        return "Maharashtra State Board (MSBSHSE)"
    if "vedanta" in nl and "college" in nl: return "Vedanta College"
    if ("oriantal" in nl or "oriental" in nl) and "sanpada" in nl:
        return "Oriental College, Sanpada"
    if "snr" in nl and "college" in nl: return "SNR Degree College"
    return n.strip()

# ── Parse multi-line key:value block ─────────────────────────────────────────
def parse_block(text, fields):
    result = {f: "" for f in fields}
    if not isinstance(text, str):
        return result
    for line in text.strip().splitlines():
        if ":" in line:
            k, _, v = line.partition(":")
            k = k.strip()
            if k in result:
                result[k] = v.strip()
    return result

EDU_FIELDS = ["Level","Field Of Study","College","Start Date","End Date",
              "Batch ID","Is Pursing","Centre Name","Is Sahi Trained","What did you learn"]
FAM_FIELDS = ["Relationship","Salutation","Name","Mobile Number","Email",
              "Gender","Age","DOB","Qualification","Average Annual Income","Occupation"]

# ── Read Excel helper ─────────────────────────────────────────────────────────
def read_excel(path, skip_empty_col=0):
    if not os.path.exists(path):
        print(f"  WARNING: {path} not found")
        return []
    t = time.time()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    raw_h = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else "" for h in raw_h]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[skip_empty_col] is None:
            continue
        r = {}
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else None
            if isinstance(v, (int, float)):
                r[h] = float(v)
            else:
                r[h] = str(v).strip() if v is not None else ""
        rows.append(r)
    wb.close()
    print(f"  {os.path.basename(path)}: {len(rows)} rows in {time.time()-t:.2f}s")
    return rows

# ── LOAD & PROCESS AT STARTUP ─────────────────────────────────────────────────
print("=" * 60)
print("ECP Impact Dashboard — Loading raw data...")

# 1. Load ECP
print("\n[1/3] Loading ECP.xlsx ...")
raw_ecp = read_excel(os.path.join(DATA_DIR, "ECP.xlsx"), skip_empty_col=0)
raw_ecp = [r for r in raw_ecp if r.get("Candidate Last Status","") in ALLOWED_STATUSES]
print(f"  After status filter: {len(raw_ecp)} candidates")

# 2. Load Attendance
print("\n[2/3] Loading Attendance_data.xlsx ...")
raw_att = read_excel(os.path.join(DATA_DIR, "Attendance_data.xlsx"), skip_empty_col=0)

# 3. Load FA
print("\n[3/3] Loading FA.xlsx ...")
raw_fa = read_excel(os.path.join(DATA_DIR, "FA.xlsx"), skip_empty_col=0)

# ── COMPUTE ATT SUMMARY per candidate ────────────────────────────────────────
print("\nComputing Attendance summary...")
att_agg = defaultdict(lambda: {"total":0,"present":0,"absent":0,"modules":set()})
for r in raw_att:
    key = r.get("Batch ID","").strip() + "|" + r.get("Candidate ID","").strip()
    a   = att_agg[key]
    a["total"] += 1
    if r.get("Attendance Status","") == "Present": a["present"] += 1
    else: a["absent"] += 1
    a["modules"].add(r.get("Module Name",""))

ATT_KEY = {}
for key, a in att_agg.items():
    pct = round(a["present"]/a["total"]*100, 2) if a["total"] else 0
    if   pct >= 90: band = "90–100% (Excellent)"
    elif pct >= 75: band = "75–89% (Good)"
    elif pct >= 50: band = "50–74% (Low)"
    elif pct  > 0:  band = "<50% (Critical)"
    else:           band = "No Attendance Data"
    ATT_KEY[key] = {"att_pct": pct, "att_band": band,
                    "total_sessions": a["total"], "present": a["present"], "absent": a["absent"]}

# ── COMPUTE FA SUMMARY per candidate ─────────────────────────────────────────
print("Computing FA summary...")
fa_agg = defaultdict(lambda: {"total":0,"passed":0,"failed":0,"attempted":0,"not_attempted":0})
for r in raw_fa:
    key = r.get("Batch ID","").strip() + "|" + r.get("Candidate ID","").strip()
    a   = fa_agg[key]
    a["total"] += 1
    s = r.get(FA_STATUS_COL,"")
    if s == "Passed":          a["passed"]       += 1
    elif s == "Failed":        a["failed"]        += 1
    elif s == "Attemped":      a["attempted"]     += 1
    elif s == "Not Attempted": a["not_attempted"] += 1

FA_KEY = {}
for key, a in fa_agg.items():
    pct = round(a["passed"]/a["total"]*100, 2) if a["total"] else 0
    if   pct >= 80: status = "Excellent (≥80%)"
    elif pct >= 60: status = "Good (60–79%)"
    elif pct  > 0:  status = "Needs Improvement (<60%)"
    else:           status = "Not Evaluated"
    FA_KEY[key] = {"fa_pct": pct, "fa_status": status,
                   "passed": a["passed"], "failed": a["failed"],
                   "attempted": a["attempted"], "not_attempted": a["not_attempted"]}

# ── BUILD MASTER CANDIDATE LIST ───────────────────────────────────────────────
print("Building master candidate list...")
CANDIDATES = []
for r in raw_ecp:
    key = r.get("Batch ID","").strip() + "|" + r.get("Candidate ID","").strip()
    att = ATT_KEY.get(key, {"att_pct":0,"att_band":"No Attendance Data","total_sessions":0,"present":0,"absent":0})
    fa  = FA_KEY.get(key,  {"fa_pct":0,"fa_status":"Not Evaluated","passed":0,"failed":0,"attempted":0,"not_attempted":0})

    # Parse Education Details
    edu = parse_block(r.get("Education Details",""), EDU_FIELDS)
    edu["College"] = normalise_college(edu.get("College",""))

    # Parse Family Details
    fam = parse_block(r.get("Family Details",""), FAM_FIELDS)

    is_cert  = r.get("Assessment Certification Status","").lower() == "certified"
    is_placed = r.get("Has Placed","").lower() == "yes" and is_cert

    CANDIDATES.append({
        # Identifiers
        "_key":       key,
        "batch_id":   r.get("Batch ID","").strip(),
        "cand_id":    r.get("Candidate ID","").strip(),
        "batch_cand": r.get("Batch ID","").strip() + " & " + r.get("Candidate ID","").strip(),

        # ECP fields
        "project":    r.get("Project Name",""),
        "sub_project":r.get("Sub Project Name",""),
        "centre_id":  r.get("Centre ID",""),
        "centre":     r.get("Centre Name",""),
        "centre_type":r.get("Centre Type",""),
        "batch_start":str(r.get("Batch Actual Start Date",""))[:10],
        "batch_end":  str(r.get("Batch Actual End Date",""))[:10],
        "name":       r.get("Candidate Name",""),
        "gender":     r.get("Candidate Gender",""),
        "phone":      r.get("Candidate Phone",""),
        "state":      r.get("Present State",""),
        "district":   r.get("Present District",""),
        "course":     r.get("Course Name",""),
        "qp":         r.get("QP Name",""),
        "edu_level":  r.get("Highest Qualification",""),
        "status":     r.get("Candidate Last Status",""),
        "assessed":   r.get("Has Assessed","").lower() == "yes",
        "cert_status":r.get("Assessment Certification Status",""),
        "is_cert":    is_cert,
        "is_placed":  is_placed,
        "emp_type":   r.get("Employment Type",""),
        "job_title":  r.get("Job Title",""),
        "company":    r.get("Placement Company Or Branch Name",""),

        # Attendance
        "att_pct":    att["att_pct"],
        "att_band":   att["att_band"],
        "total_sessions": att["total_sessions"],
        "present":    att["present"],
        "absent":     att["absent"],

        # FA
        "fa_pct":     fa["fa_pct"],
        "fa_status":  fa["fa_status"],
        "fa_passed":  fa["passed"],
        "fa_failed":  fa["failed"],

        # Education & Family (parsed)
        "college":    edu["College"],
        "edu_field":  edu["Field Of Study"],
        "fam_rel":    fam["Relationship"],
        "fam_occ":    fam["Occupation"],
        "fam_income": fam["Average Annual Income"],
    })

print(f"Master list: {len(CANDIDATES)} candidates")

# ── BUILD FILTER OPTION SETS ──────────────────────────────────────────────────
ALL_PROJECTS = sorted(set(c["project"] for c in CANDIDATES if c["project"]))
ALL_CENTRES  = sorted(set(c["centre"]  for c in CANDIDATES if c["centre"]))
ALL_BATCHES  = sorted(set(c["batch_id"] for c in CANDIDATES if c["batch_id"]))
ALL_STATUSES = sorted(set(c["status"]   for c in CANDIDATES if c["status"]))

# Project → Centres mapping for cascading
PROJ_CENTRES = defaultdict(set)
for c in CANDIDATES:
    if c["project"] and c["centre"]:
        PROJ_CENTRES[c["project"]].add(c["centre"])
PROJ_CENTRES = {k: sorted(v) for k, v in PROJ_CENTRES.items()}

# Centre → Batches mapping
CEN_BATCHES = defaultdict(set)
for c in CANDIDATES:
    if c["centre"] and c["batch_id"]:
        CEN_BATCHES[c["centre"]].add(c["batch_id"])
CEN_BATCHES = {k: sorted(v) for k, v in CEN_BATCHES.items()}

print(f"\nReady!")
print(f"  Projects: {len(ALL_PROJECTS)} | Centres: {len(ALL_CENTRES)} | Batches: {len(ALL_BATCHES)}")
print("=" * 60)

# ── FILTER HELPERS ────────────────────────────────────────────────────────────
def get_filter_args():
    return {
        "projects": request.args.getlist("project"),
        "centres":  request.args.getlist("centre"),
        "batches":  request.args.getlist("batch"),
        "statuses": request.args.getlist("status"),
    }

def apply_filters(pool, filters):
    p = filters.get("projects",[])
    c = filters.get("centres",[])
    b = filters.get("batches",[])
    s = filters.get("statuses",[])
    if p: pool = [r for r in pool if r["project"] in p]
    if c: pool = [r for r in pool if r["centre"]  in c]
    if b: pool = [r for r in pool if r["batch_id"] in b]
    if s: pool = [r for r in pool if r["status"]   in s]
    return pool

def n(v):
    try: return int(float(v or 0))
    except: return 0

def pct(x, t, d=1):
    return round(x/t*100, d) if t else 0

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

# ── Filter options (cascading) ────────────────────────────────────────────────
@app.route("/api/filters")
def api_filters():
    f = get_filter_args()
    pool = apply_filters(CANDIDATES, f)

    # Cascading: centres depend on selected projects
    proj_sel = f["projects"]
    if proj_sel:
        valid_centres = sorted(set(c for p in proj_sel for c in PROJ_CENTRES.get(p,[])))
    else:
        valid_centres = ALL_CENTRES

    # Batches depend on selected projects + centres
    cen_sel  = f["centres"]
    batch_pool = apply_filters(CANDIDATES, {"projects":proj_sel,"centres":cen_sel,"batches":[],"statuses":[]})
    valid_batches = sorted(set(c["batch_id"] for c in batch_pool if c["batch_id"]))

    return jsonify({
        "projects": ALL_PROJECTS,
        "centres":  valid_centres,
        "batches":  valid_batches,
        "statuses": ALL_STATUSES,
    })

# ── Overview / KPIs ───────────────────────────────────────────────────────────
@app.route("/api/overview")
def api_overview():
    f    = get_filter_args()
    pool = apply_filters(CANDIDATES, f)
    total = len(pool)

    assessed    = sum(1 for c in pool if c["assessed"])
    certified   = sum(1 for c in pool if c["is_cert"])
    cert_placed = sum(1 for c in pool if c["is_placed"])
    female      = sum(1 for c in pool if c["gender"] == "Female")
    att_vals    = [c["att_pct"] for c in pool if c["att_pct"] > 0]
    fa_vals     = [c["fa_pct"]  for c in pool if c["fa_pct"]  > 0]
    att_avg     = round(sum(att_vals)/len(att_vals),1) if att_vals else 0
    fa_avg      = round(sum(fa_vals)/len(fa_vals),1)   if fa_vals  else 0
    in_att      = len(att_vals)
    in_fa       = len(fa_vals)

    # Breakdowns
    status_vc = defaultdict(int)
    att_bands = defaultdict(int)
    fa_bands  = defaultdict(int)
    edu_vc    = defaultdict(int)
    state_vc  = defaultdict(int)
    course_vc = defaultdict(int)
    monthly   = defaultdict(int)

    for c in pool:
        status_vc[c["status"]] += 1
        att_bands[c["att_band"]] += 1
        fa_bands[c["fa_status"]]  += 1
        if c["edu_level"]: edu_vc[c["edu_level"]] += 1
        if c["state"]:     state_vc[c["state"]]   += 1
        if c["course"]:    course_vc[c["course"].replace("Certificate Course for ","").replace("Certification Course for ","")] += 1
        if c["batch_start"] and c["batch_start"] != "nan":
            mo = c["batch_start"][:7]
            if mo: monthly[mo] += 1

    # Sort monthly
    monthly_sorted = dict(sorted(monthly.items()))

    return jsonify({
        "kpis": {
            "total": total, "assessed": assessed,
            "certified": certified, "cert_placed": cert_placed,
            "female": female, "male": total - female,
            "att_avg": att_avg, "fa_avg": fa_avg,
            "centres": len(set(c["centre"]  for c in pool)),
            "batches":  len(set(c["batch_id"] for c in pool)),
            "assessment_rate":  pct(assessed, total),
            "certification_rate": pct(certified, total),
            "placement_rate":   pct(cert_placed, certified),
        },
        "pipeline": {
            "enrolled":   total, "in_att":     in_att,
            "in_fa":      in_fa,  "assessed":   assessed,
            "certified":  certified, "placed":  cert_placed,
        },
        "status_vc":  dict(sorted(status_vc.items(), key=lambda x:-x[1])),
        "att_bands":  dict(att_bands),
        "fa_bands":   dict(fa_bands),
        "edu_vc":     dict(sorted(edu_vc.items(), key=lambda x:-x[1])),
        "state_vc":   dict(sorted(state_vc.items(), key=lambda x:-x[1])),
        "course_vc":  dict(sorted(course_vc.items(), key=lambda x:-x[1])),
        "monthly":    monthly_sorted,
    })

# ── Centre Performance ─────────────────────────────────────────────────────────
@app.route("/api/centres")
def api_centres():
    f    = get_filter_args()
    pool = apply_filters(CANDIDATES, f)

    cen_map = defaultdict(lambda: {"project":"","total":0,"cert":0,"placed":0,
                                    "att_sum":0.0,"fa_sum":0.0,"batches":set()})
    for c in pool:
        cm = cen_map[c["centre"]]
        cm["project"] = c["project"]
        cm["total"]  += 1
        if c["is_cert"]:   cm["cert"]   += 1
        if c["is_placed"]: cm["placed"] += 1
        cm["att_sum"] += c["att_pct"]
        cm["fa_sum"]  += c["fa_pct"]
        cm["batches"].add(c["batch_id"])

    result = []
    for name, cm in cen_map.items():
        t = cm["total"]
        result.append({
            "name":          name,
            "project":       cm["project"],
            "total":         t,
            "certified":     cm["cert"],
            "placed":        cm["placed"],
            "placement_pct": pct(cm["placed"], cm["cert"]),
            "att_avg":       round(cm["att_sum"]/t, 1) if t else 0,
            "fa_avg":        round(cm["fa_sum"]/t,  1) if t else 0,
            "batches":       len(cm["batches"]),
        })
    result.sort(key=lambda x: -x["placement_pct"])
    return jsonify({"centres": result, "total": len(result)})

# ── Placement / SDG 8 ──────────────────────────────────────────────────────────
@app.route("/api/placement")
def api_placement():
    f      = get_filter_args()
    pool   = apply_filters(CANDIDATES, f)
    placed = [c for c in pool if c["is_placed"]]
    cert   = sum(1 for c in pool if c["is_cert"])
    pn     = len(placed)

    emp_vc = defaultdict(int)
    co_vc  = defaultdict(int)
    job_vc = defaultdict(int)
    fem_pl = 0
    male_pl= 0

    for c in placed:
        e = c["emp_type"] or "Unknown"
        co= c["company"]  or "Unknown"
        j = c["job_title"]or "Unknown"
        emp_vc[e] += 1
        co_vc[co] += 1
        job_vc[j] += 1
        if c["gender"] == "Female": fem_pl += 1
        else: male_pl += 1

    fem_all  = sum(1 for c in pool if c["gender"]=="Female")
    male_all = sum(1 for c in pool if c["gender"]=="Male")

    return jsonify({
        "kpis": {
            "total": len(pool), "placed": pn, "certified": cert,
            "placement_pct": pct(pn, cert),
            "fem_placed":  fem_pl, "male_placed": male_pl,
            "unique_employers": len(co_vc),
        },
        "emp_vc":        dict(emp_vc),
        "top_companies": sorted(co_vc.items(),  key=lambda x:-x[1])[:20],
        "top_jobs":      sorted(job_vc.items(),  key=lambda x:-x[1])[:10],
        "gender": {
            "fem_enrolled": fem_all, "male_enrolled": male_all,
            "fem_placed":   fem_pl,  "male_placed":   male_pl,
        },
    })

# ── Candidates (paginated) ─────────────────────────────────────────────────────
@app.route("/api/candidates")
def api_candidates():
    f      = get_filter_args()
    pool   = apply_filters(CANDIDATES, f)
    search = request.args.get("search","").lower().strip()
    gender = request.args.get("gender","")
    page   = int(request.args.get("page",1))
    per_pg = int(request.args.get("per_page",30))

    if search: pool = [c for c in pool if search in c["name"].lower() or search in c["batch_id"].lower()]
    if gender: pool = [c for c in pool if c["gender"] == gender]

    total  = len(pool)
    start  = (page-1)*per_pg
    subset = pool[start:start+per_pg]

    return jsonify({
        "candidates": [{
            "name":     c["name"],
            "id":       c["batch_cand"],
            "batch_id": c["batch_id"],
            "gender":   c["gender"],
            "centre":   c["centre"],
            "project":  c["project"],
            "status":   c["status"],
            "cert":     c["cert_status"],
            "is_cert":  c["is_cert"],
            "is_placed":c["is_placed"],
            "course":   c["course"],
            "att_pct":  c["att_pct"],
            "fa_pct":   c["fa_pct"],
            "att_band": c["att_band"],
            "fa_status":c["fa_status"],
            "state":    c["state"],
        } for c in subset],
        "total":       total,
        "page":        page,
        "per_page":    per_pg,
        "total_pages": (total+per_pg-1)//per_pg,
    })

# ── Insights: College & Family ─────────────────────────────────────────────────
@app.route("/api/insights")
def api_insights():
    f    = get_filter_args()
    pool = apply_filters(CANDIDATES, f)

    college_vc = defaultdict(int)
    rel_map    = defaultdict(lambda: {"count":0,"incomes":[],"occupations":defaultdict(int)})
    occ_map    = defaultdict(lambda: {"count":0,"incomes":[]})

    for c in pool:
        college_vc[c["college"]] += 1

        rel = c["fam_rel"] or "Unknown"
        occ = c["fam_occ"] or "Unknown"
        try: inc = float(c["fam_income"] or 0)
        except: inc = 0

        rel_map[rel]["count"] += 1
        if inc > 0: rel_map[rel]["incomes"].append(inc)
        rel_map[rel]["occupations"][occ] += 1

        occ_map[occ]["count"] += 1
        if inc > 0: occ_map[occ]["incomes"].append(inc)

    # College list
    college_list = sorted(
        [{"college":k,"count":v} for k,v in college_vc.items()],
        key=lambda x:-x["count"]
    )

    # Relationship list with top occupations
    rel_list = []
    for rel, d in sorted(rel_map.items(), key=lambda x:-x[1]["count"]):
        avg = round(sum(d["incomes"])/len(d["incomes"])) if d["incomes"] else 0
        top_occ = sorted(d["occupations"].items(), key=lambda x:-x[1])[:5]
        rel_list.append({
            "relationship":    rel,
            "count":           d["count"],
            "avg_income":      avg,
            "top_occupations": [[o,n] for o,n in top_occ],
        })

    # Occupation by count
    occ_by_count = sorted(
        [{"occupation":k,"count":d["count"]} for k,d in occ_map.items()],
        key=lambda x:-x["count"]
    )[:15]

    # Occupation by avg income (min 3 entries)
    occ_by_income = sorted(
        [{"occupation":k,
          "avg_income": round(sum(d["incomes"])/len(d["incomes"]))}
         for k,d in occ_map.items() if len(d["incomes"]) >= 3],
        key=lambda x:-x["avg_income"]
    )[:15]

    return jsonify({
        "colleges":       college_list,
        "relationships":  rel_list,
        "occ_by_count":   occ_by_count,
        "occ_by_income":  occ_by_income,
        "total_colleges": len(college_list),
        "total_relations":len(rel_list),
    })

# ── Attendance Summary table ───────────────────────────────────────────────────
@app.route("/api/summary/att")
def api_att_summary():
    f      = get_filter_args()
    pool   = apply_filters(CANDIDATES, f)
    search = request.args.get("search","").lower()
    page   = int(request.args.get("page",1))
    per_pg = int(request.args.get("per_page",25))

    rows = []
    for c in pool:
        if search and search not in c["batch_id"].lower() and search not in c["cand_id"].lower():
            continue
        rows.append({
            "Batch ID":       c["batch_id"],
            "Candidate ID":   c["cand_id"],
            "Project Name":   c["project"],
            "Sub Project Name": c["sub_project"],
            "Centre ID":      c["centre_id"],
            "Centre Name":    c["centre"],
            "Candidate Name": c["name"],
            "Total Sessions": c["total_sessions"],
            "Present":        c["present"],
            "Absent":         c["absent"],
            "Attendance %":   c["att_pct"],
            "Attendance Band":c["att_band"],
        })

    total  = len(rows)
    start  = (page-1)*per_pg
    subset = rows[start:start+per_pg]
    cols   = list(subset[0].keys()) if subset else []

    return jsonify({
        "rows": subset, "cols": cols, "total": total,
        "page": page, "total_pages": (total+per_pg-1)//per_pg,
    })

# ── FA Summary table ───────────────────────────────────────────────────────────
@app.route("/api/summary/fa")
def api_fa_summary():
    f      = get_filter_args()
    pool   = apply_filters(CANDIDATES, f)
    search = request.args.get("search","").lower()
    page   = int(request.args.get("page",1))
    per_pg = int(request.args.get("per_page",25))

    rows = []
    for c in pool:
        if search and search not in c["batch_id"].lower() and search not in c["cand_id"].lower():
            continue
        rows.append({
            "Batch ID":        c["batch_id"],
            "Candidate ID":    c["cand_id"],
            "Project Name":    c["project"],
            "Sub Project Name":c["sub_project"],
            "Centre ID":       c["centre_id"],
            "Centre Name":     c["centre"],
            "Candidate Name":  c["name"],
            "Gender":          c["gender"],
            "FA Pass Rate %":  c["fa_pct"],
            "FA Passed":       c["fa_passed"],
            "FA Failed":       c["fa_failed"],
            "FA Status":       c["fa_status"],
        })

    total  = len(rows)
    start  = (page-1)*per_pg
    subset = rows[start:start+per_pg]
    cols   = list(subset[0].keys()) if subset else []

    return jsonify({
        "rows": subset, "cols": cols, "total": total,
        "page": page, "total_pages": (total+per_pg-1)//per_pg,
    })

# ── Session Photos ─────────────────────────────────────────────────────────────
@app.route("/api/photos")
def api_photos():
    batch_ids = set(request.args.getlist("batch"))
    if not batch_ids:
        return jsonify({"photos":[],"message":"Select a batch to view photos"})

    photos = []
    for r in raw_att:
        if len(photos) >= 60: break
        if r.get("Batch ID","").strip() not in batch_ids: continue
        raw_p = r.get("Session Photos","")
        if not raw_p: continue
        urls  = re.findall(r"https://[^\s\n]+\.jpg", raw_p)
        date  = str(r.get("Session Date",""))[:10]
        mod   = str(r.get("Module Name","")).strip()[:50]
        for url in urls:
            photos.append({"url":url,"date":date,"module":mod})
            if len(photos) >= 60: break

    return jsonify({"photos":photos,"total":len(photos)})


if __name__ == "__main__":
    app.run(debug=True, port=5000)
